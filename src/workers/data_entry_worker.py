"""Data-entry worker -- generates spreadsheets from structured data.

``DataEntryWorker`` handles ``GigType.DATA_ENTRY`` orders.  It extracts the
task description, expected columns, and row expectations from the order,
sends them through the ``data_entry_task`` prompt, parses the structured
JSON response into rows, and saves the result as an ``.xlsx`` Excel file.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from src.ai.client import AIClient
from src.ai.prompts import PromptManager
from src.models.schemas import GigType, Order
from src.utils.file_handler import DeliverableManager
from src.utils.logger import get_logger
from src.workers.base import BaseWorker

log = get_logger(__name__, component="data_entry_worker")


class DataEntryWorker(BaseWorker):
    """Produce spreadsheet deliverables from structured data tasks."""

    def __init__(
        self,
        client: AIClient,
        prompts: PromptManager,
        file_manager: DeliverableManager,
    ) -> None:
        super().__init__(client, prompts, file_manager)

    # ------------------------------------------------------------------
    # Abstract property
    # ------------------------------------------------------------------

    @property
    def gig_type(self) -> GigType:
        return GigType.DATA_ENTRY

    # ------------------------------------------------------------------
    # Analysis
    # ------------------------------------------------------------------

    async def analyze(self, order: Order) -> dict[str, Any]:
        """Extract data-entry-specific parameters from the order.

        Returns
        -------
        dict
            Keys: ``description``, ``columns``, ``row_count``,
            ``data_sources``, ``additional_instructions``.
        """
        requirements_text = "\n".join(order.requirements)

        system = (
            "You are a data entry order analyst. Extract the following fields "
            "from the buyer's requirements and return them as a JSON object:\n"
            '- "description": concise description of the data entry task\n'
            '- "columns": list of expected column headers for the spreadsheet\n'
            '- "row_count": estimated number of rows needed (integer, default 10 if not specified)\n'
            '- "data_sources": list of data sources mentioned (URLs, files, etc.)\n'
            '- "output_format": desired output format description\n'
            '- "additional_instructions": any other specific instructions\n\n'
            "Return ONLY valid JSON."
        )
        user = (
            f"Order ID: {order.fiverr_order_id}\n"
            f"Requirements:\n{requirements_text}"
        )

        log.info("data_entry_worker.analyze", order_id=order.id)

        try:
            data = await self._client.complete_json(
                system=system,
                user=user,
                purpose="data_entry_analysis",
            )
        except Exception:
            log.exception("data_entry_worker.analyze_failed", order_id=order.id)
            data = {}

        return {
            "description": data.get("description", requirements_text[:200] or "Data entry task"),
            "columns": data.get("columns", []),
            "row_count": int(data.get("row_count", 10)),
            "data_sources": data.get("data_sources", []),
            "output_format": data.get("output_format", "Excel spreadsheet"),
            "additional_instructions": data.get("additional_instructions", ""),
        }

    # ------------------------------------------------------------------
    # Data parsing helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_tabular_data(
        raw_response: str,
    ) -> tuple[list[str], list[list[Any]]]:
        """Parse Claude's response into headers and rows.

        The model is instructed to return a JSON object with ``headers``
        and ``rows`` keys.  This method handles markdown fences and
        provides fallback parsing for edge cases.

        Returns
        -------
        tuple[list[str], list[list[Any]]]
            A ``(headers, rows)`` tuple.
        """
        # Strip optional markdown code fences.
        cleaned = raw_response.strip()
        if cleaned.startswith("```"):
            lines = cleaned.split("\n")
            if lines[-1].strip() == "```":
                lines = lines[1:-1]
            else:
                lines = lines[1:]
            cleaned = "\n".join(lines).strip()

        try:
            data = json.loads(cleaned)
        except json.JSONDecodeError:
            log.error(
                "data_entry_worker.parse_json_failed",
                raw_preview=raw_response[:300],
            )
            # Fallback: return raw text as a single-cell table.
            return ["Content"], [[raw_response]]

        # Expected shape: {"headers": [...], "rows": [[...], ...]}
        if isinstance(data, dict):
            headers = data.get("headers", [])
            rows = data.get("rows", data.get("data", []))
            if headers and rows:
                return headers, rows

            # Maybe the dict values themselves are the data (flat list of dicts).
            if not headers and not rows:
                # Try treating each dict entry as a row.
                for key, value in data.items():
                    if isinstance(value, list) and len(value) > 0:
                        if isinstance(value[0], dict):
                            headers = list(value[0].keys())
                            rows = [list(item.values()) for item in value]
                            return headers, rows
                        elif isinstance(value[0], list):
                            return [f"Column_{i+1}" for i in range(len(value[0]))], value

        # If data is a list of dicts, extract headers from keys.
        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                rows = [list(item.values()) for item in data]
                return headers, rows
            elif isinstance(data[0], list):
                return [f"Column_{i+1}" for i in range(len(data[0]))], data

        # Last resort.
        return ["Content"], [[str(data)]]

    # ------------------------------------------------------------------
    # Execution
    # ------------------------------------------------------------------

    async def execute(
        self, order: Order, analysis: dict[str, Any]
    ) -> list[Path]:
        """Generate structured data and save it as an Excel spreadsheet.

        Returns
        -------
        list[Path]
            Single-element list with the path to the ``.xlsx`` file.
        """
        log.info(
            "data_entry_worker.execute",
            order_id=order.id,
            row_count=analysis["row_count"],
            column_count=len(analysis["columns"]),
        )

        system = self._prompts.get_system("data_entry_task")

        # Build the input data description from requirements.
        input_data = "\n".join(order.requirements)
        if analysis["data_sources"]:
            input_data += "\n\nData sources: " + ", ".join(analysis["data_sources"])

        user = self._prompts.get_user(
            "data_entry_task",
            description=analysis["description"],
            input_data=input_data or "Generate sample data based on the description",
            output_format=(
                f"Return a JSON object with 'headers' (list of column names) "
                f"and 'rows' (list of lists). "
                f"Expected columns: {', '.join(analysis['columns']) if analysis['columns'] else 'determine appropriate columns'}. "
                f"Target approximately {analysis['row_count']} rows."
            ),
            additional_instructions=analysis["additional_instructions"] or "None",
        )

        # Scale token budget with expected data volume.
        estimated_tokens = max(analysis["row_count"] * 100, 2048)
        max_tokens = min(estimated_tokens, 8192)

        try:
            raw_response = await self._client.complete(
                system=system,
                user=user,
                max_tokens=max_tokens,
                temperature=0.2,
                purpose="data_entry_execution",
            )
        except Exception:
            log.exception(
                "data_entry_worker.execute_api_failed", order_id=order.id
            )
            raise

        headers, rows = self._parse_tabular_data(raw_response)

        # Determine filename.
        safe_desc = (
            analysis["description"]
            .replace(" ", "_")
            .replace("/", "_")
            .replace("\\", "_")[:40]
        )
        filename = f"{safe_desc}.xlsx"

        path = self._file_manager.save_xlsx(
            order_id=order.id,
            filename=filename,
            data=rows,
            headers=headers,
        )

        log.info(
            "data_entry_worker.deliverable_saved",
            order_id=order.id,
            row_count=len(rows),
            column_count=len(headers),
            path=str(path),
        )
        return [path]

    # ------------------------------------------------------------------
    # Revision
    # ------------------------------------------------------------------

    async def revise(
        self,
        order: Order,
        feedback: str,
        original_paths: list[Path],
    ) -> list[Path]:
        """Revise the spreadsheet based on buyer feedback.

        Loads the original data from the Excel file, sends it along with
        the revision feedback to Claude, and saves the updated spreadsheet.

        Returns
        -------
        list[Path]
            Single-element list with the path to the revised ``.xlsx`` file.
        """
        log.info(
            "data_entry_worker.revise",
            order_id=order.id,
            feedback_length=len(feedback),
        )

        # Read original data from the Excel file.
        original_data_text = ""
        if original_paths:
            try:
                from openpyxl import load_workbook

                wb = load_workbook(str(original_paths[0]))
                ws = wb.active
                if ws is not None:
                    all_rows: list[list[Any]] = []
                    for row in ws.iter_rows(values_only=True):
                        all_rows.append([str(cell) if cell is not None else "" for cell in row])
                    if all_rows:
                        # First row is headers.
                        headers_line = " | ".join(all_rows[0])
                        data_lines = [" | ".join(r) for r in all_rows[1:]]
                        original_data_text = (
                            f"Headers: {headers_line}\n"
                            f"Data ({len(data_lines)} rows):\n"
                            + "\n".join(data_lines)
                        )
            except Exception:
                log.warning(
                    "data_entry_worker.revise_read_failed",
                    path=str(original_paths[0]),
                )
                original_data_text = "[Original data could not be read]"

        system = self._prompts.get_system("revision_task")

        revision_context = (
            "The original deliverable is an Excel spreadsheet. "
            "Return the revised data as a JSON object with 'headers' "
            "(list of column names) and 'rows' (list of lists)."
        )
        user = self._prompts.get_user(
            "revision_task",
            gig_type="data_entry",
            original_content=original_data_text,
            revision_feedback=feedback,
            additional_context=revision_context,
        )

        try:
            raw_response = await self._client.complete(
                system=system,
                user=user,
                max_tokens=8192,
                temperature=0.2,
                purpose="data_entry_revision",
            )
        except Exception:
            log.exception(
                "data_entry_worker.revise_api_failed", order_id=order.id
            )
            raise

        headers, rows = self._parse_tabular_data(raw_response)

        # Save revised file.
        revision_num = order.revision_count + 1
        base_name = original_paths[0].stem if original_paths else "data"
        filename = f"{base_name}_rev{revision_num}.xlsx"

        path = self._file_manager.save_xlsx(
            order_id=order.id,
            filename=filename,
            data=rows,
            headers=headers,
        )

        log.info(
            "data_entry_worker.revision_saved",
            order_id=order.id,
            revision=revision_num,
            row_count=len(rows),
            path=str(path),
        )
        return [path]
